from __future__ import annotations

import hashlib
import re
import shutil
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
import requests
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.engine import Engine

from .config import PROJECT_ROOT


RIDERSHIP_RESOURCES = {
    "historical": {
        "package_url": "https://ckan0.cf.opendata.inter.prod-toronto.ca/api/3/action/package_show?id=ttc-ridership-analysis",
        "resource_name": "1985-2019 Analysis of ridership",
        "default_filename": "1985-2019-analysis-of-ridership.xlsx",
        "notes": "Historical matrix counts are published in thousands and multiplied by 1,000 during ingest.",
    },
    "surface": {
        "package_url": "https://ckan0.cf.opendata.inter.prod-toronto.ca/api/3/action/package_show?id=ttc-ridership-all-day-weekday-for-surface-routes",
        "resource_name": "ranking-surface-routes",
        "default_filename": "ranking-surface-routes.xlsx",
        "notes": "Surface ridership sample date is read from workbook header when available.",
    },
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _resource(kind: str) -> dict[str, Any]:
    config = RIDERSHIP_RESOURCES[kind]
    response = requests.get(config["package_url"], timeout=60)
    response.raise_for_status()
    package = response.json()["result"]
    for resource in package.get("resources", []):
        if resource.get("name") == config["resource_name"]:
            return resource
    raise RuntimeError(f"Could not find CKAN resource {config['resource_name']!r}.")


def download_ridership_files(force: bool = False) -> dict[str, Path]:
    data_dir = PROJECT_ROOT / "data" / "ridership"
    data_dir.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}
    for kind, config in RIDERSHIP_RESOURCES.items():
        resource = _resource(kind)
        path = data_dir / config["default_filename"]
        if force or not path.exists():
            with requests.get(resource["url"], stream=True, timeout=180) as response:
                response.raise_for_status()
                with path.open("wb") as handle:
                    for chunk in response.iter_content(1024 * 1024):
                        if chunk:
                            handle.write(chunk)
        paths[kind] = path
    return paths


def archive_ridership_file(path: Path, kind: str) -> Path:
    archive_dir = PROJECT_ROOT / "data" / "ridership" / "archives"
    archive_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    digest = sha256_file(path)
    archive_path = archive_dir / f"{kind}_{stamp}_{digest[:12]}{path.suffix}"
    if not archive_path.exists():
        shutil.copy2(path, archive_path)
    return archive_path


def parse_historical_matrix(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    year_columns: list[tuple[int, int]] = []
    for column in range(3, ws.max_column + 1):
        raw = ws.cell(6, column).value
        match = re.search(r"\d{4}", str(raw or ""))
        if match:
            year_columns.append((column, int(match.group(0))))

    records: list[dict[str, Any]] = []
    rider_type: str | None = None
    for row in range(7, ws.max_row + 1):
        section = str(ws.cell(row, 1).value or "").strip().upper()
        media = str(ws.cell(row, 2).value or "").strip()

        if section in {"WHERE", "WHEN"}:
            break
        if section == "WHO" and media:
            rider_type = media
            continue
        if media in {"SENIOR/YOUTHS", "CHILDREN"}:
            rider_type = media
            continue
        if media == "DAY/VIST./OTHER":
            rider_type = "OTHER"
        if not media or rider_type is None:
            continue
        if "SUB-TOTAL" in media.upper() or "SYSTEM TOTAL" in media.upper():
            continue

        clean_media = re.sub(r"\s+", " ", media).strip()
        for column, year in year_columns:
            raw_count = ws.cell(row, column).value
            if raw_count in (None, "N/A", ""):
                continue
            count = int(float(raw_count) * 1000)
            records.append(
                {
                    "year": year,
                    "media": clean_media,
                    "rider_type": rider_type,
                    "count": count,
                    "source_file": path.name,
                }
            )
    return pd.DataFrame.from_records(records)


def parse_surface_routes(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    sample_date = date(2016, 12, 31)
    for row in range(1, min(ws.max_row, 10) + 1):
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row, column).value
            if isinstance(value, str):
                match = re.search(r"([A-Za-z]+)\s+(\d{1,2}),\s+(\d{4})", value)
                if match:
                    sample_date = datetime.strptime(match.group(0), "%B %d, %Y").date()

    rows: list[dict[str, Any]] = []
    for row in range(8, ws.max_row + 1):
        rank = ws.cell(row, 1).value
        route_id = ws.cell(row, 2).value
        route_name = ws.cell(row, 3).value
        riders = ws.cell(row, 4).value
        if route_id in (None, "") or route_name in (None, ""):
            continue
        numeric_riders = pd.to_numeric(riders, errors="coerce")
        rows.append(
            {
                "route_id": str(route_id).strip(),
                "route_name": str(route_name).strip(),
                "rank": int(rank) if rank not in (None, "") else None,
                "all_day_riders": int(numeric_riders) if pd.notna(numeric_riders) else None,
                "sample_date": sample_date,
                "source_file": path.name,
            }
        )
    return pd.DataFrame.from_records(rows)


def upsert_dataframe(engine: Engine, table: str, rows: pd.DataFrame) -> int:
    if rows.empty:
        return 0
    temp_table = f"staging_{table}"
    with engine.begin() as conn:
        conn.execute(text(f"DROP TABLE IF EXISTS {temp_table}"))
        conn.execute(text(f"CREATE UNLOGGED TABLE {temp_table} (LIKE {table} INCLUDING DEFAULTS)"))
    rows.to_sql(temp_table, engine, if_exists="append", index=False, method="multi")
    if table == "ridership_matrix":
        sql = f"""
            INSERT INTO ridership_matrix (year, media, rider_type, count, source_file)
            SELECT year, media, rider_type, count, source_file FROM {temp_table}
            ON CONFLICT (year, media, rider_type)
            DO UPDATE SET
                count = EXCLUDED.count,
                source_file = EXCLUDED.source_file,
                ingested_at = now()
        """
    elif table == "surface_route_ridership":
        sql = f"""
            INSERT INTO surface_route_ridership (
                route_id, route_name, rank, all_day_riders, sample_date, source_file
            )
            SELECT route_id, route_name, rank, all_day_riders, sample_date, source_file FROM {temp_table}
            ON CONFLICT (route_id)
            DO UPDATE SET
                route_name = EXCLUDED.route_name,
                rank = EXCLUDED.rank,
                all_day_riders = EXCLUDED.all_day_riders,
                sample_date = EXCLUDED.sample_date,
                source_file = EXCLUDED.source_file,
                ingested_at = now()
        """
    else:
        raise ValueError(f"Unsupported table: {table}")
    with engine.begin() as conn:
        conn.execute(text(sql))
        conn.execute(text(f"DROP TABLE IF EXISTS {temp_table}"))
    return len(rows)


def record_ingestion_run(
    engine: Engine,
    kind: str,
    source_path: Path,
    archive_path: Path,
    row_count: int,
) -> int:
    resource = _resource(kind)
    config = RIDERSHIP_RESOURCES[kind]
    with engine.begin() as conn:
        return int(
            conn.execute(
                text(
                    """
                    INSERT INTO ridership_ingestion_runs (
                        source_name,
                        source_url,
                        portal_resource_id,
                        portal_last_modified,
                        file_sha256,
                        archive_path,
                        row_count,
                        notes
                    )
                    VALUES (
                        :source_name,
                        :source_url,
                        :portal_resource_id,
                        NULLIF(:portal_last_modified, '')::timestamptz,
                        :file_sha256,
                        :archive_path,
                        :row_count,
                        :notes
                    )
                    ON CONFLICT (file_sha256)
                    DO UPDATE SET
                        ingested_at = now(),
                        archive_path = EXCLUDED.archive_path,
                        row_count = EXCLUDED.row_count
                    RETURNING ridership_ingestion_id
                    """
                ),
                {
                    "source_name": resource.get("name", config["resource_name"]),
                    "source_url": resource.get("url", ""),
                    "portal_resource_id": resource.get("id", ""),
                    "portal_last_modified": resource.get("last_modified")
                    or resource.get("metadata_modified")
                    or "",
                    "file_sha256": sha256_file(source_path),
                    "archive_path": str(archive_path),
                    "row_count": row_count,
                    "notes": config["notes"],
                },
            ).scalar_one()
        )


def validate_ridership(engine: Engine) -> dict[str, int]:
    sql = {
        "ridership_matrix_rows": "SELECT COUNT(*) FROM ridership_matrix",
        "surface_route_ridership_rows": "SELECT COUNT(*) FROM surface_route_ridership",
        "negative_matrix_counts": "SELECT COUNT(*) FROM ridership_matrix WHERE count < 0",
        "negative_surface_counts": "SELECT COUNT(*) FROM surface_route_ridership WHERE all_day_riders < 0",
        "surface_routes_missing_gtfs_match": """
            SELECT COUNT(*)
            FROM surface_route_ridership srr
            LEFT JOIN routes r ON r.route_id = srr.route_id
            WHERE r.route_id IS NULL
        """,
    }
    with engine.connect() as conn:
        return {name: int(conn.execute(text(query)).scalar_one()) for name, query in sql.items()}


def ingest_ridership(engine: Engine, force_download: bool = False) -> dict[str, Any]:
    paths = download_ridership_files(force=force_download)
    historical = parse_historical_matrix(paths["historical"])
    surface = parse_surface_routes(paths["surface"])

    historical_count = upsert_dataframe(engine, "ridership_matrix", historical)
    surface_count = upsert_dataframe(engine, "surface_route_ridership", surface)

    historical_archive = archive_ridership_file(paths["historical"], "historical")
    surface_archive = archive_ridership_file(paths["surface"], "surface")
    run_ids = {
        "historical": record_ingestion_run(
            engine, "historical", paths["historical"], historical_archive, historical_count
        ),
        "surface": record_ingestion_run(
            engine, "surface", paths["surface"], surface_archive, surface_count
        ),
    }
    return {
        "run_ids": run_ids,
        "row_counts": {
            "ridership_matrix": historical_count,
            "surface_route_ridership": surface_count,
        },
        "validation": validate_ridership(engine),
    }
